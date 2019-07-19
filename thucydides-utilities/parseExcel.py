from openpyxl import load_workbook
import sqlite3
import xml.etree.ElementTree as ET
import random
import re
import json
import subprocess

import simplifyUnicode
import parseDefinition
import utils


TEXT_LOCATION = "input/ThucydidesPelopGrWithPunct.xml"
DB_LOCATION = "input/GreekMorphologyThucEurPlato.db"

# escape single quotes in a string
def sqliteEscapeString(s):
    return s.replace("'", "''")

# create a class name associated with a semantic group
def createSemanticGroupBadgeClass(s):
    s = s.lower()
    s = s.replace(" ", "-")
    s += "-badge"
    return s

def sanitizeCompoundString(s):
    if s == None:
        return ""
    item = s
    item = item.replace("-", "")
    item = item.replace("?", "")
    trimmed = item.strip()
    return trimmed

# Remove acute accents
def noAcute(s):
    s = re.sub(r'[άἄἅ]', 'α', s)
    s = re.sub(r'[έἔἕ]', 'ε', s)
    s = re.sub(r'[ήἤἥ]', 'η', s)
    s = re.sub(r'[ίἴἵ]', 'ι', s)
    s = re.sub(r'[όὄὅ]', 'ο', s)
    s = re.sub(r'[ύὔὕ]', 'υ', s)
    s = re.sub(r'[ώὤὥ]', 'ω', s)
    return s

# Remove numbers
def noNums(s):
    s = re.sub(r'\d', '', s)
    return s

HTML_TAB = "  "

def indentByTab(s):
    s = s.replace("\n", "\n%s" % HTML_TAB)
    s = HTML_TAB + s
    return s

# functions for prepping the text
def prepSection(bookNum, chapterNum, sectionNum, preppedSectionStorage, preppedChapterStorage, preppedTexts):
    sectionKey = "%d.%d.%d" % (bookNum, chapterNum, sectionNum)
    preppedSectionText = ""
    currentContext = None

    # determine context of each token
    contexts = []
    for token in preppedSectionStorage:
        split = token.split("@")
        if len(split) > 1:
            tokenContext = split[1]
            if tokenContext != currentContext:
                # contexts.append(int(tokenContext))
                if tokenContext == "0":
                    contexts.append(0)
                elif tokenContext == "1":
                    contexts.append(1)
                elif tokenContext == "2":
                    contexts.append(2)
                elif tokenContext == "3":
                    contexts.append(3)
                else:
                    print("BAD CONTEXT, THIS IS BAD")
                    contexts.append(-1)
        else:
            contexts.append(-1)
    # if punctuation is surrounded by context of a certain type, it is
    # part of that context group
    for i, context in enumerate(contexts):
        if context == -1:
            if (i == 0):
                contexts[i] = contexts[1]
            elif (i == len(contexts)-1):
                contexts[i] = contexts[-2]
            else:
                # find the next non-punctuation
                nextIndex = i+1
                while nextIndex < len(contexts) and contexts[nextIndex] == -1:
                    nextIndex += 1

                # if we reached the end of the sentence or the context is the same,
                # this punctuation is part of the same group.
                if nextIndex >= len(contexts) or contexts[i-1] == contexts[nextIndex]:
                    contexts[i] = contexts[i-1]

    # store where context groups start and end
    contextStarts = []
    contextEnds = []
    for i, context in enumerate(contexts):
        # start
        if (i == 0):
            contextStarts.append(context)
        else:
            if context != -1 and context != contexts[i-1]:
                contextStarts.append(context)
            else:
                contextStarts.append(-1)

        # end
        if (i == len(contexts)-1):
            contextEnds.append(1)
        else:
            if context != -1 and context != contexts[i+1]:
                contextEnds.append(1)
            else:
                contextEnds.append(0)



    for i, token in enumerate(preppedSectionStorage):
        split = token.split("@")

        tokenContext = contexts[i]
        if contextStarts[i] >= 0:
            context = contextStarts[i]
            contextClass = ""

            contextClass = "context-%d" % context

            # move space outside the contex
            if len(split[0]) > 0 and split[0][0] == " ":
                preppedSectionText += " "
                split[0] = split[0][1:]

            preppedSectionText += "<span class=\"%s\">" % contextClass
            currentContext = tokenContext

        if len(split) > 1:
            preppedSectionText += split[0] + "@" + "@".join(split[2:])
        else:
            preppedSectionText += token

        # end the context if necessary
        if contextEnds[i] > 0:
            preppedSectionText += "</span>"

    #preppedSectionText = "".join(preppedSectionStorage)
    preppedTexts[sectionKey] = "<p>\n%s%s\n</p>" % (HTML_TAB, preppedSectionText)
    # add section to corresponding chapter
    preppedChapterStorage.append("<p>")
    preppedChapterStorage.append("%s<span class=\"sectionLabel\">%d</span>" % (HTML_TAB, sectionNum))
    preppedChapterStorage.append(indentByTab(preppedSectionText))
    preppedChapterStorage.append("</p>")

def prepChapter(bookNum, chapterNum, preppedChapterStorage, preppedBookStorage, preppedTexts):
    chapterKey = "%d.%d" % (bookNum, chapterNum)
    preppedChapterText = "\n".join(preppedChapterStorage)
    preppedTexts[chapterKey] = preppedChapterText
    # add chapter to corresponding book
    preppedBookStorage.append("<div>")
    preppedBookStorage.append("%s<h3>Chapter %d</h3>" % (HTML_TAB, chapterNum))
    preppedBookStorage.append(indentByTab(preppedChapterText))
    preppedBookStorage.append("</div>")

def prepBook(bookNum, preppedBookStorage, preppedTexts):
    bookKey = "%d" % bookNum
    preppedTexts[bookKey] = "\n".join(preppedBookStorage)

# true if matching Betant info should not be used for this lemma
def ignoreBetant(lemma):
    return (lemma == "Δῆλος" or lemma == "Ἕλος" or lemma == "Ἐπίκουρος" or
            lemma == "Εὔβουλος" or lemma == "Ἠπειρωτικός" or lemma == "Ἵππαρχος"
            or lemma == "Ἄκρα" or lemma == "Κῦρος" or lemma == "Μητρόπολις"
            or lemma == "Ὅμηρος" or lemma == "οὗ2" or lemma == "Παράλιος"
            or lemma == "Πόλις" or lemma == "Πόντος" or lemma == "Πρόξενος" or
            lemma == "προσδέω" or lemma == "Τέως" or lemma == "ἄπειρος2" or
            lemma == "Εὐφήμως" or lemma == "ἔφοδος" or lemma == "ἔφοδος2" or
            lemma == "θύω2" or lemma == "καταδέω2" or lemma == "κύριος2" or
            lemma == "περίπλοος" or lemma == "πυρά" or lemma == "χερσόνησος"
            or lemma == "δέω2")

# list of semantic groups
semanticGroups = {}


# mapping from semantic group to index
semanticIndices = {}


# function for recognizing a semantic group; returns true if there is one
def findSemanticGroup(item):
    if (item == ""):
        return False, -1

    lowerItem = item.strip().lower()

    # Make conversions between different things if necessary
    while lowerItem in semanticConversions:
        lowerItem = semanticConversions[lowerItem]

    if lowerItem in semanticIndices:
        return (True, semanticIndices[lowerItem])
    else:
        return (False, -1)

# given an array of items split by comma, find semantic groups that could be in it.
def findSemanticGroups(pieces):
    if (len(pieces) == 0):
        return (True, [], "")

    for i in range(len(pieces)):
        match, matchIndex = findSemanticGroup(",".join(pieces[:i+1]))
        if match:
            valid, found, error = findSemanticGroups(pieces[i+1:])
            found.append(matchIndex)
            return (valid, found, error)

    return (False, [], ",".join(pieces))

# function for extracting multiple semantic groups
def extractSemanticGroups(val):
    if (val == ""):
        return []

    str = sanitizeCompoundString(val)

    groups = re.split(r',', str)
    # valid, found, error = findSemanticGroups(splt)
    #
    # if not(valid):
    #     print("Semantic Group Issues: '%s', '%s'" % (val, error))


    # if ((len(found) > 1) or (len(found) == 1 and error)):
    #     print("Found Multiple '%s': %s" % (val, ", ".join(map(lambda x: "%d" % x, sorted(found)))))
    finalGroups = []
    for group in groups:
        group = group.strip()
        finalGroups.append(group)
        if (not(group in semanticGroups)):
            semanticGroups[group] = group;
    return sorted(finalGroups)


# conversions for stem types
stemConversions = {

}

# Conversions for compounds
compoundConversions = {
    "ἀντι": "ἀντί",
    "αντι": "ἀντί",
    "ἐπι": "ἐπί",
    "ἀπο": "ἀπό",
    "άπό": "ἀπό",
    "αντί": "ἀντί",
    "προ": "πρό",
    "προς": "πρός",
    "συν": "σύν",
    "έκ": "ἐκ",
    "εκ": "ἐκ",
}


TAB = "    "

ERROR_LOG = []

betant = utils.getContent("results/betant.json", True)
betantUsed = {}
betantMissing = []
for key in betant:
    betantUsed[key] = []

class Lemma(object):
    def __init__(self, row):
        super().__init__()

        self.lemma = self.extract_lemma(row)
        self.setSearchableLemmas()

        self.errorLog = []

        self.shortDef = self.extract_shortDef(row)

        self.partOfSpeech = self.extract_partOfSpeech(row)

        aa, bb, cc, dd = self.extract_longDefinition(row)
        self.hasLongDefinition = aa
        self.longDefinitionRaw = bb
        self.longDefinition = cc
        self.instanceMeanings = dd

        self.semanticGroup = self.extract_semanticGroup(row)
        self.stemType = self.extract_stemType(row)
        self.compoundParts = self.extract_compoundParts(row)
        self.frequency = self.extract_frequency(row)

        aa, bb, cc, dd = self.extract_keyPassage(row)
        self.hasKeyPassage = aa
        self.keyPassageLocation = bb
        self.keyPassageText = cc
        self.keyPassageTranslation = dd

        aa, bb, cc, dd = self.extract_illustration(row)
        self.hasIllustration = aa
        self.illustrationLink = bb
        self.illustrationAlt = cc
        self.illustrationCaption = dd

        self.bibliographyText = self.extract_bibliographyText(row)


    # functions for extracting values from row
    def extract_default(self, val):
        res = ""
        if not(val == None):
            res = val
        return res

    def extract_lemma(self, row):
        val = self.extract_default(row[1].value)
        if (re.search(",", val)):
            val = val.split(", ")[0]
        return val

    def extract_shortDef(self, row):
        shortDef = re.sub(r'"', r'\\"', self.extract_default(row[5].value))
        if (row[3].value != None):
            shortDef = self.extract_default(row[3].value)
        return shortDef

    def extract_longDefinition(self, row):
        self.betantLongDefinition = "false"
        hasLongDef = "false"
        longDefRaw = ""
        longDef = "[]"
        meanings = {}

        articleFilename = "input/articles/%s.txt" % self.lemma

        myLem = noNums(self.lemma)
        if (utils.fileExists(articleFilename)):
            print("exists " + self.lemma)
            val = utils.getContent(articleFilename, False)
            hasLongDef = "true"
            longDefRaw = val

            longDefList, meanings = parseDefinition.createLongDefObjects(longDefRaw)

            longDef = json.dumps(longDefList, sort_keys=True)
        elif ((self.lemma[-1] == "2") and (self.betacodeLemma+"2") in betant):
            # default to various betant possibilities
            hasLongDef = "true"
            raw, ld, meanings = betant[self.betacodeLemma+"2"]
            longDefRaw = raw;
            for i in range(len(ld)):
                if ("lem" in ld[i]["text"]):
                    ld[i]["text"].pop("lem")
            longDef = json.dumps(ld, sort_keys=True)
            betantUsed[self.betacodeLemma+"2"].append(self.lemma)
            self.betantLongDefinition = "true"
        elif ((self.lemma[-1] == "3") and (self.betacodeLemma+"3") in betant):
            hasLongDef = "true"
            raw, ld, meanings = betant[self.betacodeLemma+"3"]
            longDefRaw = raw;
            for i in range(len(ld)):
                if ("lem" in ld[i]["text"]):
                    ld[i]["text"].pop("lem")
            longDef = json.dumps(ld, sort_keys=True)
            betantUsed[self.betacodeLemma+"3"].append(self.lemma)
            self.betantLongDefinition = "true"
        elif (self.betacodeLemma in betant):
            if not(ignoreBetant(self.lemma)):
                hasLongDef = "true"
                raw, ld, meanings = betant[self.betacodeLemma]
                longDefRaw = raw;
                for i in range(len(ld)):
                    if ("lem" in ld[i]["text"]):
                        ld[i]["text"].pop("lem")
                longDef = json.dumps(ld, sort_keys=True)
                betantUsed[self.betacodeLemma].append(self.lemma)
                self.betantLongDefinition = "true"
        elif (self.lemma in betant):
            hasLongDef = "true"
            raw, ld, meanings = betant[self.lemma]
            longDefRaw = raw;
            for i in range(len(ld)):
                if ("lem" in ld[i]["text"]):
                    ld[i]["text"].pop("lem")
            longDef = json.dumps(ld, sort_keys=True)
            betantUsed[self.lemma].append(self.lemma)
            self.betantLongDefinition = "true"
        else:
            betantMissing.append(self.lemma)
            #print(self.betacodeLemma, "not in Betant")

        return hasLongDef, longDefRaw, longDef, meanings

    def extract_partOfSpeech(self, row):
        return re.sub(r'"', r'\\"', self.extract_default(row[19].value))

    def extract_semanticGroup(self, row):
        val = row[9].value
        if not(val == None):
            newItems = extractSemanticGroups(val)
        else:
            newItems = []

        return newItems

    def extract_stemType(self, row):
        # possible we extract multiple stems
        val = row[6].value
        if not(val == None):
            items = re.split(r'[,\s]', val)
            newItems = []
            for item in items:
                item = sanitizeCompoundString(item)
                if not(item == ""):
                    if item in stemConversions:
                        newItems.append(stemConversions[item])
                    else:
                        newItems.append(item)
        else:
            newItems = []

        return newItems

    def extract_compoundParts(self, row):
        val = row[2].value
        if not(val == None):
            items = re.split(r'[,\s]', val)
            newItems = []
            for item in items:
                item = sanitizeCompoundString(item)
                if not(item == ""):
                    if item in compoundConversions:
                        newItems.append(compoundConversions[item])
                    else:
                        newItems.append(item)
        else:
            newItems = []

        return newItems

    def extract_frequency(self, row):
        # Currently Perseus Only
        allFreq = self.extract_default(row[21].value)
        try:
            allFreq = int(allFreq)
        except ValueError:
            if not(allFreq == ""):
                ERROR_LOG.append("%s is not an integer!" % allFreq)
            allFreq = 0
        allFreq = 0
        obj = {
            "all": allFreq,
        }
        return obj

    # key passages are in the definitions now so this is a dinosaur
    def extract_keyPassage(self, row):
        val = self.extract_default(row[4].value)
        #m = re.match(r'(\d+\.\d+\.\d+)\s+(.*)', val)

        has = "false"
        loc = ""
        text = ""
        transl = ""
        return has, loc, text, transl

        # if m:
        #     has = "true"
        #     loc = m.group(1)
        #
        #     m2 = re.match(r'^(.*?)(".*")?$', m.group(2))
        #     text = m2.group(1)
        #     if (m2.group(2) != None):
        #         transl = m2.group(2)
        #
        #     # #print(val)
        #     # print(self.lemma)
        #     # print(m.group(2))
        #     # print(m2.group(1))
        #     # print(m2.group(2))
        #     # print("----")
        # return has, loc, text, transl

    def extract_illustration(self, row):
        has = "false"
        link = ""
        alt = ""
        caption = ""

        imageFilenameJPG = "%s.jpg" % self.lemma
        imageFilenameGIF = "%s.gif" % self.lemma
        imageFilenamePNG = "%s.png" % self.lemma
        imageFilenameTxt = "input/illustrations/%s.txt" % self.lemma

        someImage = False

        if (utils.fileExists("input/illustrations/" + imageFilenameJPG)):
            someImage = True
            imageFilename = imageFilenameJPG
            extension = ".jpg"
        elif (utils.fileExists("input/illustrations/" + imageFilenameGIF)):
            someImage = True
            imageFilename = imageFilenameGIF
            extension = ".gif"
        elif (utils.fileExists("input/illustrations/" + imageFilenamePNG)):
            someImage = True
            imageFilename = imageFilenamePNG
            extension = ".png"

        if someImage:
            has = "true"
            link = "%s%s" % (self.betacodeUnaccentedLemma, extension)

            # copy the image into a new folder, with a name that isn't
            # in unicode.
            runStr = "cp input/illustrations/%s results/illustrations/%s" % (imageFilename, link)
            subprocess.run(runStr, shell=True)

            if utils.fileExists(imageFilenameTxt):
                caption = utils.getContent(imageFilenameTxt, False)
                alt = caption

        return has, link, alt, caption

    def extract_illustrationLink(self, row):
        return "test.png"

    def extract_illustrationAlt(self, row):
        return "alt text"

    def extract_illustrationCaption(self, row):
        return "[Caption]"

    def extract_bibliographyText(self, row):
        btext = ""
        val = self.extract_default(row[11].value)
        if val != "":
            btext = val
        return btext

    #========================================================

    #
    def setSearchableLemmas(self):
        self.searchLemma = simplifyUnicode.searchVersion(self.lemma)
        self.unaccentedLemma = simplifyUnicode.unaccented(self.lemma)
        self.betacodeLemma = simplifyUnicode.betacode(self.lemma)
        self.betacodeUnaccentedLemma = simplifyUnicode.unaccentedBetacode(self.lemma)
        self.latinApproximationLemma = simplifyUnicode.latinApproximation(self.lemma)

    # get the key for this lemma
    def getKey(self):
        return self.lemma

    # get full typescript entry
    def getFullEntryTS(self):
        s = []
        s.append("  {")
        s.append("%stoken: \"%s\"," % (TAB, self.lemma))
        s.append("%ssearch: [" % (TAB))
        s.append("%s  \"%s\"," % (TAB, self.lemma))
        s.append("%s  \"%s\"," % (TAB, self.searchLemma))
        s.append("%s  \"%s\"," % (TAB, self.unaccentedLemma))
        s.append("%s  \"%s\"," % (TAB, self.betacodeLemma))
        s.append("%s  \"%s\"," % (TAB, self.betacodeUnaccentedLemma))
        s.append("%s  \"%s\"," % (TAB, self.latinApproximationLemma))
        s.append("%s]," % (TAB))
        s.append("%sshortDef: \"%s\"," % (TAB, self.shortDef))
        s.append("%shasLongDefinition: %s," % (TAB, self.hasLongDefinition))
        s.append("%sbetantLongDefinition: %s," % (TAB, self.betantLongDefinition))
        s.append("%slongDefinitionRaw: \"%s\"," % (TAB, self.longDefinitionRaw))
        s.append("%slongDefinition: \"%s\"," % (TAB, self.longDefinition))
        s.append("%spartOfSpeech: \"%s\"," % (TAB, self.partOfSpeech))
        s.append("%ssemanticGroup: [\"%s\"]," % (TAB, "\", \"".join(self.stemType)))
        s.append("%sstemType: [\"%s\"]," % (TAB, "\", \"".join(self.stemType)))
        s.append("%scompoundParts: [\"%s\"]," % (TAB, "\", \"".join(self.compoundParts)))
        s.append("%sfrequency: {" % (TAB))
        s.append("%s  all: %d," % (TAB, self.frequency["all"]))
        s.append("%s}," % (TAB))
        s.append("%shasKeyPassage: %s," % (TAB, self.hasKeyPassage))
        s.append("%skeyPassageLocation: \"%s\"," % (TAB, self.keyPassageLocation))
        s.append("%skeyPassageText: \"%s\"," % (TAB, self.keyPassageText))
        s.append("%skeyPassageTranslation: \"%s\"," % (TAB, self.keyPassageTranslation))
        s.append("%shasIllustration: %s," % (TAB, self.hasIllustration))
        s.append("%sillustrationLink: \"%s\"," % (TAB, self.illustrationLink))
        s.append("%sillustrationAlt: \"%s\"," % (TAB, self.illustrationAlt))
        s.append("%sillustrationCaption: \"%s\"," % (TAB, self.illustrationCaption))
        s.append("%sbibliographyText: \"%s\"," % (TAB, self.bibliographyText))
        s.append("  }")
        return "\n".join(s)

    # get short typescript entry
    def getShortEntryTS(self):
        s = []
        s.append("  {")
        s.append("%stoken: \"%s\"," % (TAB, self.lemma))
        s.append("%ssearch: [" % (TAB))
        s.append("%s  \"%s\"," % (TAB, self.lemma))
        s.append("%s  \"%s\"," % (TAB, self.searchLemma))
        s.append("%s  \"%s\"," % (TAB, self.unaccentedLemma))
        s.append("%s  \"%s\"," % (TAB, self.betacodeLemma))
        s.append("%s  \"%s\"," % (TAB, self.betacodeUnaccentedLemma))
        s.append("%s  \"%s\"," % (TAB, self.latinApproximationLemma))
        s.append("%s]," % (TAB))
        s.append("%sshortDef: \"%s\"," % (TAB, self.shortDef))
        s.append("  }")
        return "\n".join(s)

    # get the text for a command to insert this lemma into a table
    def getSQLiteInsertCommand(self, tableName):
        s = []
        s.append("INSERT INTO %s VALUES (" % tableName)
        s.append("\'%s\', " % sqliteEscapeString(self.lemma))
        s.append("\'%s\', " % sqliteEscapeString(self.shortDef))

        # search information - now in a different database
        # s.append("\'%s\', " % sqliteEscapeString(self.searchLemma))
        # s.append("\'%s\', " % sqliteEscapeString(self.unaccentedLemma))
        # s.append("\'%s\', " % sqliteEscapeString(self.betacodeLemma))
        # s.append("\'%s\', " % sqliteEscapeString(self.betacodeUnaccentedLemma))
        # s.append("\'%s\', " % sqliteEscapeString(self.latinApproximationLemma))

        if (self.hasLongDefinition == "true"):
            s.append("1, ")
        else:
            s.append("0, ")
        # If this is a betant long definition, it is old and the author is betant (2)
        if (self.betantLongDefinition == "true"):
            s.append("1, 2, ")
        else:
            # Otherwise it isn't old and the author is lexeis admin.
            s.append("0, 1, ")

        # No custom author
        s.append("\'\', ")
        s.append("\'%s\', " % sqliteEscapeString(self.longDefinitionRaw))
        s.append("\'%s\', " % sqliteEscapeString(self.longDefinition))

        s.append("\'%s\', " % sqliteEscapeString(self.partOfSpeech))
        s.append("\'%s\', " % sqliteEscapeString(json.dumps(self.semanticGroup)))
        s.append("\'%s\', " % sqliteEscapeString(json.dumps(self.stemType)))
        s.append("\'%s\', " % sqliteEscapeString(json.dumps(self.compoundParts)))

        s.append("%d, " % (self.frequency["all"]))

        if (self.hasKeyPassage == "true"):
            s.append("1, ")
        else:
            s.append("0, ")
        s.append("\'%s\', " % sqliteEscapeString(self.keyPassageLocation))
        s.append("\'%s\', " % sqliteEscapeString(self.keyPassageText))
        s.append("\'%s\', " % sqliteEscapeString(self.keyPassageTranslation))

        if (self.hasIllustration == "true"):
            s.append("1, ")
        else:
            s.append("0, ")
        s.append("\'%s\', " % sqliteEscapeString(self.illustrationLink))
        s.append("\'%s\', " % sqliteEscapeString(self.illustrationAlt))
        s.append("\'%s\', " % sqliteEscapeString(self.illustrationCaption))

        s.append("\'%s\'" % sqliteEscapeString(self.bibliographyText))
        s.append(")")


        return "".join(s)

    # get the text for a command to insert this lemma's searchable info
    def getSQLiteSearchCommands(self, tableName):
        commands = []
        searches = [
            self.searchLemma,
            self.unaccentedLemma,
            self.betacodeLemma,
            self.betacodeUnaccentedLemma,
            self.latinApproximationLemma
        ]

        # συν ξυν
        if ("σύν" in self.compoundParts):
            searches.append(re.sub(r'σ(ύ|υ)', r'ξ\1', self.searchLemma))
            searches.append(re.sub(r'συ', r'ξυ', self.unaccentedLemma))
            searches.append(re.sub(r'su', r'cu', self.betacodeLemma))
            searches.append(re.sub(r'su', r'cu', self.betacodeUnaccentedLemma))
            searches.append(re.sub(r'su', r'xu', self.latinApproximationLemma))

            # matches = re.findall(r'(σ(ύ|υ))', self.lemma)
            # if len(matches) > 1:
            #     print(self.lemma)
            #     print(matches)
            #     print("---")

        # final sigma
        if (self.lemma[-1] == "ς"):
            searches.append(self.searchLemma[:-1] + "σ")
            searches.append(self.unaccentedLemma[:-1] + "σ")


        for search in searches:
            s = []
            s.append("INSERT INTO %s VALUES (" % tableName)
            # Extra blank for potential alias
            s.append("\'%s\', \'\'," % sqliteEscapeString(self.lemma))

            # search information
            s.append("\'%s\'" % sqliteEscapeString(search))
            s.append(")")
            res = "".join(s)
            commands.append(res)

        return commands

    def __str__(self):
        return "Lemma: %s" % self.lemma

# ==============================================================================
# ================================ Excel Loading  ==============================
# ==============================================================================
def getLemmaInfo():

    # remove the illustrations folder and make a new one
    runStr = "rm -r results/illustrations"
    subprocess.run(runStr, shell=True)
    runStr = "mkdir results/illustrations"
    subprocess.run(runStr, shell=True)

    workbook = load_workbook('input/dictionaryExcel.xlsx', read_only=True)
    main_sheet = workbook['Master Layout']

    count = 0
    countFull = 0
    lemmas = {}

    # print errors about lemmas from perseus's text analysis not in our dictionary
    ERROR_LOG.append("Frequencies in dictionary that aren't integers:")
    ERROR_LOG.append("----------------")

    for row in main_sheet:
        # skip header row
        if count == 0:
            count += 1
            continue
        # check if row has information; for now that means there is something in
        # the second column, which has lemma information
        if not(row[1].value == None):
            if not(row[1].value == "omit"):
                countFull += 1
                myLemma = Lemma(row)
                lemmas[myLemma.getKey()] = myLemma

        count += 1

        #if count > 3:
        #    break
    ERROR_LOG.append("================")
    return lemmas

# ==============================================================================
# ================================ Excel Loading  ==============================
# ==============================================================================

# populate the sqlite database
def createSQLiteDatabase(lemmas):

    stemTypes = {}
    allCompounds = {}
    alphaCombos = {}
    for key in lemmas:
        l = lemmas[key]

        # keep track of the possible first two letters
        uL = l.unaccentedLemma
        firstChar = uL[0].upper()
        if not(firstChar in alphaCombos):
            alphaCombos[firstChar] = {}

        # some lemmas have only a single character
        if (len(uL) == 1):
            secondChar = "_"
            twoChars = firstChar + secondChar
            alphaText = firstChar
        else:
            secondChar = uL[1]
            if (secondChar == 'ς'):
                secondChar = 'σ'
            twoChars = firstChar + secondChar
            alphaText = twoChars

        if not(twoChars in alphaCombos[firstChar]):
            link = "/wordList/%s/%s" % (firstChar, secondChar)
            alphaCombos[firstChar][twoChars] = {
                "text": alphaText,
                "link": link,
                "active": False
            }

        sts = l.stemType
        for st in sts:
            if not(st in stemTypes):
                stemTypes[st] = 1
            else:
                stemTypes[st] += 1

        compounds = l.compoundParts
        for c in compounds:
            if not(c in allCompounds):
                allCompounds[c] = 1
            else:
                allCompounds[c] += 1

    alphaCombos2 = {}
    for key in alphaCombos:
        alphaCombos2[key] = sorted(list(alphaCombos[key].values()), key=lambda ac: ac["text"])
    # save list of letter combos
    f = open('results/alphaCombos.ts', 'w')
    f.write("export const ALPHA_COMBOS = " + json.dumps(alphaCombos2, indent=2, sort_keys=True))
    f.close()

    # create a list storing the word contexts
    contexts = []
    firstContextWord = True

    rowIndex = 0
    for row in main_sheet:
        rowIndex += 1
        if rowIndex == 1:
            continue

        use = row[0].value
        book = row[1].value

        if use == 1 and not(book == None):
            chapter_start = row[2].value
            section_start = row[3].value
            word_start = row[4].value
            chapter_end = row[5].value
            section_end = row[6].value
            word_end = row[7].value
            contextName = row[9].value
            if contextName != None:
                contextName = contextName.lower()

            if word_start[0] == "“":
                word_start = word_start[1:]

            if contextName == "narrative" or contextName == "Narrative":
                context = 0
            elif contextName == "direct speech" or contextName == "direct":
                context = 1
            elif contextName == "indirect speech" or contextName == "indirext speech" or contextName == "indirect" or contextName == "indirect statement":
                context = 2
            elif contextName == "authorial" or contextName == "year-end authorial":
                context = 3
            # if contextName == "book 1":
            #     context = 0
            # elif contextName == "book 2":
            #     context = 1
            # elif contextName == "book 3":
            #     context = 2
            # elif contextName == "book 4":
            #     context = 3
            # elif contextName == "book 5":
            #     context = 4
            # elif contextName == "book 6":
            #     context = 5
            # elif contextName == "book 7":
            #     context = 6
            # elif contextName == "book 8":
            #     context = 7
            else:
                print("EMPTY/BAD CONTEXT at row %d" % rowIndex)
                context = 0
            obj = {
                "book": book,
                "start": [chapter_start, section_start, word_start.lower().strip()],
                "end": [chapter_end, section_end, word_end.lower().strip()],
                "context": context,
                "row": rowIndex,
            }

            contexts.append(obj)

    # initialize context information
    contextIndex = -1
    currentContext = contexts[0]

    # text information
    tree = ET.parse(TEXT_LOCATION)
    root = tree.getroot()
    text = root[1][0]

    # open thucyides database
    conn = sqlite3.connect(DB_LOCATION)
    c = conn.cursor()

    instanceInformationInputs = []
    textStorageInputs = []
    sectionContextInputs = []

    # object for holding sections compactly
    sectionsList = {}

    # full list of sections
    allSections = {}

    # pre-preppared text
    preppedTexts = {}
    preppedSectionStorage = []
    preppedChapterStorage = []
    preppedBookStorage = []
    isFollowingStartPunct = False

    # lemmas in perseus chicago but not ours
    noMatchLemmas = {}
    # list of lemmas by sections
    sectionLemmas = {}

    lastTokenInfo = []

    bookNum = 0
    totalTokenCount = 0
    for book in text:
        bookNum += 1
        chapterNum = 0
        sectionNum = 0
        wordIndex = 0
        trueIndex = 0


        # store a book's chapters
        chapterInfo = {}
        # store section
        sectionInfo = {}

        print("Handling book: %d" % bookNum)
        wordArray = []
        for el in book[0]:

            # most children are just words, but some sections contain
            # lyric on quote bits that have their own children which each
            # need to be handled separately.
            if (el.tag == "l" or el.tag == "p"): # lyric and quotes
                wordArray.extend(el)
            else:
                wordArray.append(el)

            # else:
            #     sectionCode = "%d.%d.%d" % (bookNum, chapterNum, sectionNum)
            #     print(sectionCode, child.tag)

        for tokenIndex, child in enumerate(wordArray):
            if (child.tag == "milestone"):
                milestoneType = child.attrib['unit']
                if (milestoneType == "chapter"):
                    # clean up previous chapter
                    if (chapterNum != 0):
                        prepSection(bookNum, chapterNum, sectionNum, preppedSectionStorage, preppedChapterStorage, preppedTexts)
                        preppedSectionStorage = []

                        prepChapter(bookNum, chapterNum, preppedChapterStorage, preppedBookStorage, preppedTexts)
                        preppedChapterStorage = []

                        if (sectionNum == 1):
                            prevSection = ""
                        else:
                            prevSection = str(sectionNum - 1)
                        nextSection = ""
                        sectionInfo[sectionNum] = {
                            "_prev": prevSection,
                            "_next": nextSection
                        }
                        sectionInfo["_last"] = sectionNum
                        sectionInfo["_first"] = 1

                        if (chapterNum == 1):
                            prevChapter = ""
                        else:
                            prevChapter = str(chapterNum - 1)
                        nextChapter = str(chapterNum + 1)
                        sectionInfo["_prev"] = prevChapter
                        sectionInfo["_next"] = nextChapter
                        chapterInfo[chapterNum] = sectionInfo

                    # set up this chapter
                    sectionInfo = {}
                    chapterNum += 1
                    sectionNum = 0

                    isFollowingStartPunct = False

                elif (milestoneType == "section"):
                    # clean up previous section
                    if (sectionNum != 0):
                        prepSection(bookNum, chapterNum, sectionNum, preppedSectionStorage, preppedChapterStorage, preppedTexts)
                        preppedSectionStorage = []


                        if (sectionNum == 1):
                            prevSection = ""
                        else:
                            prevSection = str(sectionNum - 1)
                        nextSection = str(sectionNum + 1)
                        sectionInfo[sectionNum] = {
                            "_prev": prevSection,
                            "_next": nextSection
                        }

                    # set up this section
                    sectionNum += 1
                    wordIndex = 0
                    trueIndex = 0

                    sectionCode = "%d.%d.%d" % (bookNum, chapterNum, sectionNum)
                    allSections[sectionCode] = "1"

                    isFollowingStartPunct = False
            elif (child.tag == "w"):
                trueIndex += 1
                sectionCode = "%d.%d.%d" % (bookNum, chapterNum, sectionNum)
                if not(sectionCode in sectionLemmas):
                    sectionLemmas[sectionCode] = {}
                if ('id' in child.attrib):
                    wordIndex += 1
                    myIndex = int(child.attrib['id'])

                    # select parse that was chosen by an authority, if no
                    # authority, choose the one with the highest probability
                    bestProb = 0
                    rows = c.execute('SELECT * FROM parses WHERE tokenid=%d' % myIndex)
                    for row in rows:
                        auth = row[5]
                        prob = row[7]
                        if not(auth == None):
                            lemIndex = row[2]
                            bestProb = prob
                            break
                        if prob > bestProb:
                            lemIndex = row[2]
                            bestProb = prob

                    if (lemIndex == None):
                        lemma = ""
                    else:
                        for row2 in c.execute('SELECT * FROM Lexicon WHERE lexid=%d' % lemIndex):
                            lemma = row2[3]

                    # TODO
                    if (lemma == "Θρασύβουλος"):
                        lemma = "Θρασυβούλος"
                    elif (lemma == "πλώϊμος"):
                        lemma = "πλώιμος"

                    if contextIndex < len(contexts):
                        if contextIndex + 1 < len(contexts):
                            nextStart = contexts[contextIndex+1]["start"]
                            end = contexts[contextIndex]["end"]

                            # print(bookNum, currentContext["book"])
                            # print(chapterNum, end[0])
                            # print(sectionNum, end[1])
                            # print(child.text,  end[2])
                            # print("----")

                            # check whether this token matches the list of files
                            tokenMatches = False
                            if(child.text.lower() == nextStart[2]):
                                tokenMatches = True
                            else:
                                split = nextStart[2].split(" ")
                                if len(split) > 1:
                                    # print(split)
                                    # print(wordArray[tokenIndex:tokenIndex+len(split)])
                                    # print(list(map(lambda x: x.text.lower(), wordArray[tokenIndex:tokenIndex+len(split)])))
                                    match = True
                                    for i in range(len(split)):
                                        if (wordArray[tokenIndex+i].text == None or split[i] != wordArray[tokenIndex+i].text.lower()):
                                            match = False
                                    tokenMatches = match

                            # update context
                            if bookNum == contexts[contextIndex+1]["book"] and chapterNum == nextStart[0] and sectionNum == nextStart[1] and tokenMatches:
                                if (contextIndex > 0):
                                    lastWas = "%d %d %d %s" % (lastTokenInfo[0], lastTokenInfo[1], lastTokenInfo[2], lastTokenInfo[3])
                                    lastShould =  "%d %d %d %s" % (contexts[contextIndex]["book"], end[0], end[1], end[2])
                                    if (lastWas != lastShould):
                                        print("Row %d" % contexts[contextIndex]["row"])
                                        print("Last was:      ", lastWas)
                                        print("Last should be:", lastShould)
                                        print("---")
                                contextIndex += 1
                                if contextIndex < len(contexts):
                                    currentContext = contexts[contextIndex]
                                # print("Moving to row %d" % (contextIndex+1))
                                # print(contexts[contextIndex])
                            elif bookNum > currentContext["book"] or (bookNum == currentContext["book"] and (chapterNum > end[0] or (chapterNum == end[0] and sectionNum > end[1]))):
                                print("We have:", bookNum, chapterNum, sectionNum, child.text.lower())
                                print("File says:", currentContext["book"], nextStart[0], nextStart[1], nextStart[2])
                                print("Passed end of row %d" % (contextIndex+2))
                                contextIndex += 1
                                if contextIndex < len(contexts):
                                    currentContext = contexts[contextIndex]


                        sectionContext = currentContext["context"]
                    else:
                        sectionContext = random.randint(0, 3)
                    sectionContextInputs.append("INSERT INTO section_contexts VALUES (%d, %d, %d, %d, %d, %d, %d)" % (myIndex, bookNum, chapterNum, sectionNum, wordIndex, trueIndex, sectionContext))

                    lastTokenInfo = [bookNum, chapterNum, sectionNum, child.text.lower()]

                    # add up frequencies
                    if (lemma in lemmas):
                        lemmas[lemma].frequency["all"] += 1
                        if sectionCode in lemmas[lemma].instanceMeanings:
                            lemmaMeaning = lemmas[lemma].instanceMeanings[sectionCode]
                        else:
                            lemmaMeaning = ""
                        sectionLemmas[sectionCode][lemma] = True
                    else:
                        if (lemma in noMatchLemmas):
                            noMatchLemmas[lemma].append(sectionCode)
                        else:
                            noMatchLemmas[lemma] = [sectionCode]

                    instanceInformationInputs.append("INSERT INTO instance_information VALUES (%d, '%s', %d, %d, %d, %d, %d, '%s', '%s', '%s', %d)" % (myIndex, child.text, bookNum, chapterNum, sectionNum, wordIndex, trueIndex, sectionCode, lemma, lemmaMeaning, sectionContext))
                else:
                    myIndex = -1
                    lemmaMeaning = ""
                    lemma = ""

                # add token to prepared text
                if (lemma != ""):
                    routerLink = lemma #"#/entry/%s" % lemma
                    if (lemmaMeaning != ""):
                        routerLink += ";%s" % lemmaMeaning#";meaning=%s" % lemmaMeaning

                    linkText = child.text
                    # capitalize when necessary
                    if (len(preppedSectionStorage) == 0 or (len(preppedSectionStorage) == 1 and isFollowingStartPunct)):
                        linkText = linkText.title()
                    #central = "<a class=\"%s tokenLink\" href=\"%s\">%s</a>" % (contextClass, routerLink, linkText)

                    # =====
                    # add/remove token number
                    # =====
                    central = "@%d@%s@%s@" % (sectionContext, routerLink, linkText)

                    # central = "@%d@%s@%s (%s)@" % (sectionContext, routerLink, linkText, myIndex)

                else:
                    central = "%s" % (child.text)

                isStartPunct = (myIndex == -1) and (child.text == "(" or child.text == "<" or child.text == "“" or child.text == "[")
                isEndPunct = (myIndex == -1) and not(child.text == "†" or isStartPunct)
                if (len(preppedSectionStorage) == 0 or isEndPunct or isFollowingStartPunct):
                    spaceBefore = ""
                else:
                    spaceBefore = " "

                wordHTML = "%s%s" % (spaceBefore, central)
                preppedSectionStorage.append(wordHTML)

                # insert into text_storage
                # sequence_index, token_index, token, book, chapter, section, true_word_index
                textStorageInputs.append("INSERT INTO text_storage VALUES (%d, %d, '%s', %d, %d, %d, %d, %d)" % (totalTokenCount, myIndex, child.text, bookNum, chapterNum, sectionNum, wordIndex, trueIndex))
                totalTokenCount += 1

                if (isStartPunct):
                    isFollowingStartPunct = True
                else:
                    isFollowingStartPunct = False

        if (sectionNum == 1):
            prevSection = ""
        else:
            prevSection = str(sectionNum - 1)
        nextSection = "" # this is the last section in the chapter
        sectionInfo[sectionNum] = {
            "_prev": prevSection,
            "_next": nextSection
        }
        sectionInfo["_last"] = sectionNum
        sectionInfo["_first"] = 1

        if (chapterNum == 1):
            prevChapter = ""
        else:
            prevChapter = str(chapterNum - 1)
        nextChapter = "" # this is the last chapter in the book
        sectionInfo["_prev"] = prevChapter
        sectionInfo["_next"] = nextChapter
        chapterInfo[chapterNum] = sectionInfo
        chapterInfo["_last"] = chapterNum
        chapterInfo["_first"] = 1

        if (bookNum == 1):
            prevBook = ""
        else:
            prevBook = str(bookNum - 1)
        nextBook = str(bookNum + 1)
        chapterInfo["_prev"] = prevBook
        chapterInfo["_next"] = nextBook

        sectionsList[bookNum] = chapterInfo


        # fill in prepared text
        prepSection(bookNum, chapterNum, sectionNum, preppedSectionStorage, preppedChapterStorage, preppedTexts)
        preppedSectionStorage = []

        prepChapter(bookNum, chapterNum, preppedChapterStorage, preppedBookStorage, preppedTexts)
        preppedChapterStorage = []

        prepBook(bookNum, preppedBookStorage, preppedTexts)
        preppedBookStorage = []

    sectionsList["_last"] = bookNum
    sectionsList["_first"] = 1
    sectionsList[bookNum]["_next"] = ""
    # close database
    conn.close()


    eimiMap = {
        "ἄπειμι2": "ἀπέρχομαι",
        "δίειμι": "διέρχομαι",
        "διέξειμι": "διεξέρχομαι",
        "εἶμι": "ἔρχομαι",
        "εἴσειμι": "εἰσέρχομαι",
        "ἔξειμι": "ἐξέρχομαι",
        "ἐπάνειμι": "ἐπανέρχομαι",
        "ἔπειμι2": "ἐπέρχομαι",
        "ἐπέξειμι": "ἐπεξέρχομαι",
        "κάτειμι": "κατέρχομαι",
        "μέτειμι2": "μετέρχομαι",
        "πάρειμι2": "παρέρχομαι",
        "περίειμι2": "περιέρχομαι",
        "προέξειμι2": "προεξέρχομαι",
        "πρόειμι": "προέρχομαι",
        "πρόσειμι2": "προσέρχομαι",
        "σύνειμι2": "συνέρχομαι",
        "συνέξειμι": "συνεξέρχομαι",
        "ἱερόν": "ἱερός", # technially not eimi, but same deal
        "ὅπῃ": "ὅπη",
        "πρεσβευτής": "πρέσβυς",
        "πῶς": "πως",
        "στερίσκω": "στερέω",
        "φημί": "φάσκω",
        "Χερσόνησος": "χερσόνησος",
        "χρή": "χρεών",
    }

    # store "bad" references that are recognized
    accounted = {
        "ἀγαθός": ["3.90.1"],
        "ἀγαπάω": ["6.18.4"],
        "ἀγγέλλω": ["8.50.1"],
        "ἀγρυπνία": ["3.104.4"],
        "ἀγωγή": ["4.29.1", "6.29.3"],
        "ἀγών": ["7.59.2"], # referent to another word
        "ἀδεής": ["8.89.2"], # refers to place it isn’t
        "ἀθυμέω": ["7.55.1"], # says “delete”
        "αἱρέω": ["8.92.2"], # reference
        "αἰτία": ["8.68.2"], # reference to place it isn’t
        "ἀκούω": ["6.8.4"], # reference
        "ἄκων2": ["7.57.9"], # we have ekontas
        "ἄλλος": ["8.35.4", "3.104.4"], # for first, no idea what’s going on here, for second can’t resolve variant reading
        "ἀλλότριος": ["4.44.3"], # they’re saying don’t include
        "ἅμα": ["7.1.2"], # saying its deleted here
        "ἀμαχεί": ["4.73.3"], # we have a separate lemma for this occurrence
        "ἀμύνω": ["7.40.2", "7.6.1"], # first deleted, second is a reference
        "ἀμφότερος": ["8.25.4"], # reference
        "ἀναγκάζω": ["8.95.4"], # reference
        "ἀναγκαῖος": ["1.70.2", "5.99.1"], # we have these under other words
        "ἀναγορεύω": ["8.67.2"], # we have different reading
        "ἀναδέω": ["4.48.3"], # variant reading
        "ἀναιρέω": ["1.13.6"], # variant reading
        "ἀναπαύω": ["6.99.2"], # variant reading
        "ἀναπλέω": ["6.42.1"], # reference
        "ἀναχωρέω": ["3.115.6"], # variant reading
        "ἀνέχω": ["7.34.2"], # variant reading
        "ἀνήρ": ["3.9.1"], # variant reading
        "ἀνθρώπειος": ["1.22.4"], # variant reading
        "ἀνίημι": ["4.24.4"], # reference
        "ἀνταλλάσσω": ["8.82.1"], # variant reading
        "ἀντεπάγω": ["4.25.1"], # we have two words
        "ἀντεπέξειμι": ["4.131.1"], # we have as separate word
        "ἀντικαθίζομαι": ["5.30.4", "5.6.3"], # we have separate words
        "ἀντιλαμβάνω": ["4.14.2"], # variant
        "ἀπαγγέλλω": ["1.91.2", "8.51.3"], # variant reading
        "ἅπαξ": ["5.85.1"], # we have separate entry for esapac
        "ἀπατάω": ["2.4.1"], # variant reading (ecapataw)
        "ἄπειμι2": ["1.24.5", "4.42.3"], # variant reading, not from this word (JR)
        "ἀπέχω": ["2.81.1"], # variant reading
        "ἀποδείκνυμι": ["1.77.6"], # variant reading
        "ἀποδέχομαι": ["5.83.3"], # variant reading
        "ἀπολείπω": ["7.75.4", "6.102.2"], # variant readings (text has u(po-)
        "ἀποπειράομαι": ["7.12.5"], # variant reading
        "ἄπορος": ["1.99.3"], # variant reading
        "ἀποστέλλω": ["6.2.5"], # variant reading (text has ana-stellw)
        "ἀποτείχισμα": ["7.43.1"], # variant reading
        "ἀποτρέπω": ["5.75.2", "8.108.1"], # variant readings
        "ἀποχωρέω": ["2.46.2", "2.79.6", "4.107.1", "7.70.8"], # variant readings
        "ἀργύρειος": ["6.91.7"], # variant reading
        "ἀρέσκω": ["2.72.1"], # variant reading
        "ἀρχή": ["6.20.4"], # variant reading
        "ἄρχω": ["3.92.6"], # variant reading
        "ἀσαφής": ["4.86.4"], # variant reading
        "ἀσθενής": ["7.60.2"], # variant reading
        "αὐτοκράτωρ": ["6.18.5"], # variant reading
        "αὐτονομία": ["8.64.5"], # variant reading
        "ἀφίημι": ["2.78.2"], # variant reading
        "ἀφορμή": ["6.90.3"], # variant reading
        "ἄχρηστος": ["2.78.3"], # variant reading
        "βέβαιος": ["4.31.2"], # reference to variant reading
        "βοηθέω": ["8.55.1"], # reference
        "βουλεύω": ["8.92.5"], # reference
        "γῆ": ["2.56.1"], # vr
        "γλῶσσα": ["3.58.1"], # reference
        "δεινός": ["8.46.2"], # variant reading (tad' einai instead of ta deinai)
        "δεξιός": ["1.24.1", "2.98.2", "3.24.1", "3.95.1", "3.106.1", "7.1.1", "2.81.3"], # we list these under the noun
        "δεύτερος": ["3.49.2"], # variant reading
        "δῆμος": ["2.65.4"], # reference
        "διαδίδωμι": ["1.76.2"], # reference
        "διασῴζω": ["1.129.3", "1.110.1"], # references
        "διαφέρω": ["3.39.5"], # variant reading
        "διαφθείρω": ["1.25.2"], # variant reading
        "διαχράομαι": ["1.126.11"], # variant reading
        "διέχω": ["8.95.3"], # reference
        "δοκέω": ["3.11.6", "5.16.3", "8.1.1"], # no idea, reference, no idea
        "ἐγγίγνομαι": ["2.22.2", "7.68.1"], # variant reading
        "ἐγείρω": ["7.51.1"], # variant reading
        "ἐγκαταλείπω": ["4.8.9", "4.39.2", "7.24.2"], # old reading
        "ἔγκλημα": ["1.39.3"], # variant reading
        "ἐγχειρέω": ["1.128.2"], # variant reading (epi-)
        "εἴκοσι": ["6.1.2"], # no idea where this came from
        "εἰσβάλλω": ["2.79.6"], # reference
        "εἰσβολή": ["8.31.3"], # variant, we have prosbolh
        "εἴωθα": ["1.134.4"], # variant reading
        "ἕκαστος": ["2.39.2", "4.64.3"], # variant reading, reference
        "ἑκάτερος": ["7.78.6"], # variant reading
        "ἐκεῖθεν": ["2.84.5"], # variant reading, we have ekeinwn
        "ἐκεῖσε": ["6.77.1"], # reference
        "ἐκπλέω": ["7.37.2"], # variant reading
        "ἐκπολεμέω": ["8.57.1"], # variant reading
        "ἕκτος": ["3.107.3", "5.19.1", "8.33.1"], # we have these under e(/kth
        "ἐμβάλλω": ["1.134.4"], # we have emellhsan
        "ἐμβιβάζω": ["1.53.1"], # we have esbibazw
        "ἐμπίμπλημι": ["3.82.8"], # vr
        "ἐμπίπτω": ["2.48.2"], # we have eispiptw here
        "ἔνθα": ["6.32.2"], # listed under separate word
        "ἐνοικέω": ["6.2.2"], # vr
        "ἔξειμι": ["2.21.2", "2.75.1", "3.108.1", "7.47.3", "1.46.4"], # epec- x3, ep-, variant
        "ἔξεστι": ["5.47.2", "8.18.2"], # vr
        "ἐξουσία": ["1.16.1"], # vr
        "ἐπαγωγή": ["7.34.2"], # reference
        "ἔπειμι2": ["1.72.3", "4.128.1", "4.131.2"], # para-, reference x2
        "ἐπέκεινα": ["6.63.2", "7.58.1"], #vr (ep' ekeina
        "ἐπερωτάω": ["1.25.1", "3.92.5", "8.29.1"], # under epeiromai
        "ἐπέχω": ["3.45.4"], # we have this under episxw
        "ἐπιβολή": ["1.93.6", "3.45.5"], # reference, vr
        "ἐπιγίγνομαι": ["1.126.8"], # vr en-
        "ἐπιδιώκω": ["8.34.1"], # vr no epi
        "ἐπικρατέω": ["7.42.4", "1.30.2"], # vr , ref
        "ἐπιλέγω": ["1.67.5"], # listed under epeipon
        "ἐπιπίπτω": ["4.4.1"], # vr, en- instead of epi-
        "ἐπιπλέω": ["8.79.2"], # prosplew
        "ἐπίπλοος2": ["6.32.3", "6.33.1", "8.86.5"], # we have epiploon
        "ἐπισπάω": ["4.9.2"], # we have eisbiazomai
        "ἐπιστρέφω": ["1.61.4"], # we have ἐπὶ Στρέψαν
        "ἐπιτελέω": ["1.108.3", "7.2.4"], # vr we have api-
        "ἐπιχειρέω": ["2.89.4"], # vr en-
        "ἑπτακόσιοι": ["6.96.3"], # reference
        "ἐρωτάω": ["1.90.5", "4.40.2", "7.10.1"], # under eromai x2, ephrwta
        "ἕτερος": ["3.64.3"], # listed under mhdeteros
        "ἐτήσιος": ["2.80.5"], # reference
        "ἔτι": ["5.16.3"], # ref
        "εὐπραξία": ["3.39.4"], # we list this under eupragia
        "εὐπρεπής": ["1.38.4"], # ref
        "Εὔφημος": ["3.104.5"], # no idea what is going on here
        "ἔφοδος3": ["4.129.4", "7.44.3", "7.51.2"], # listed under efodos2 x2, vr esodou
        "ἐφορμέω": ["3.31.1", "6.49.4"], # listed under eformaw
        "ἐφορμίζω": ["6.49.4"], # listed under eformaw
        "ἔφορμος2": ["3.76.1"], # listed under eformos1
        "ἔχω": ["3.26.1"], # not sure what the deal is
        "ζεῦγμα": ["7.30.2"], # reference
        "ἡγέομαι": ["8.2.1"], # not included in our text
        "ἥκω": ["1.18.2"], # vr
        "ἠπειρωτικός": ["3.94.3"], # listed under capital form
        "θεός": ["5.77.4"], # not sure about this but it doesn't really matter
        "ἱκνέομαι": ["5.40.1"], #vr, we have hkon
        "ἴσχω": ["2.91.1"], # vr, sxousai from exw
        "καθά": ["4.118.11"], # vr kaq' a(/
        "καθαιρέω": ["1.121.4"], # vr
        "καθέζομαι": ["3.104.4"], # variant reading;
        "καθίζω": ["1.136.3"], # vr kaqezomai
        "καινός": ["3.30.4"], # vr kenon
        "κατασκευάζω": ["2.85.1"], # vr para-
        "κάτειμι": ["5.7.5"], # vr an-
        "κατίσχω": ["4.42.2", "4.54.1", "4.57.3", "6.23.2", "8.23.3", "4.42.3"], # listed under katexw
        "κινδυνεύω": ["8.24.5"], # vr -cunkinduneuw
        "κλείω": ["4.8.7", "7.56.1"], # listed under klhzw
        "κοινός": ["5.79.1"], # reference
        "κομίζω": ["1.109.3"], # vr ana-
        "κομπόω": ["6.17.5"], # vr compew
        "κρατέω": ["1.30.3"], # epi-
        "κτῆσις": ["1.18.1"], # vr ktisin
        "κωλύω": ["4.32.3"], # vr
        "λέγω": ["4.20.4", "2.10.3", "8.14.2"], # leontwn for legontwn?, sunlegw, vr  (Rusten)
        "μακρός": ["7.13.2", "3.13.2", "3.55.1", "3.13.5"], # listed under makran x3, per JR, Betant parsed this incorrectly
        "μέσος": ["6.2.5", "3.82.8"], # under mesh
        "νεότης": ["4.80.3"], # vr skaiothta
        "νομίζω": ["2.94.1"], # vr
        "νόμος": ["5.70.1"], # vr o(mou= for vo/mou
        "οἰκέω": ["5.116.4", "8.108.4", "2.68.7", "6.37.2", "1.12.3"], # references to different words (wkisan, katoikew), vr x3 (Rusten)
        "οἴκησις": ["6.4.3"], # ref
        "οἰκίζω": ["1.8.1"], # vr
        "οἷος": ["4.118.4", "2.5.4"], # not here, perhaps referring to o(/saper?; we have separate singleton oi(-a
        "ὀκτώ": ["8.104.2"], # not here; ogdohkonta or ebdomhkonta taking its place?
        "ὀλιγαρχία": ["8.90.1"], # vr  ὁμολογίαν
        "ὁμολογέω": ["4.119.1"], # reference
        "ὅμορος": ["3.114.1"], # ref
        "ὁπωσοῦν": ["6.56.3"], # ref
        "ὀργάω": ["2.21.3"], # vr wrmhto
        "ὁρμέω": ["4.75.2"], # ref
        "οὐδείς": ["2.51.2"], # vr oud' en
        "οὔκουν": ["2.43.1"], # vr as ou)/k ou)=n
        "οὔπω": ["5.15.2"], # vr oupws
        "οὕτως": ["1.39.3"], # vr
        "παράλογος": ["1.140.1"], # ref
        "παρασκευάζω": ["1.142.3"], # vr kata-
        "παρασκευή": ["4.52.3"], # vra paraskuazw
        "παρέχω": ["7.36.3"], # ref
        "πάρειμι": ["2.10.3"], # vr
        "παρήκω": ["4.36.2"], # ref
        "παρίστημι": ["4.133.1"], # vr paresxhkos
        "πᾶς": ["1.123.1", "1.143.1", "2.39.2", "2.92.6", "5.47.3", "5.47.4", "5.47.7", "1.41.2", # vr apas
         "2.84.3", "6.62.1", "3.11.3"], # missing
        "πεντήκοντα": ["4.13.2"], # vr τεσσαράκοντα
        "πέραν": ["3.91.3"], # vr Γραϊκῆς for πέραν ghs
        "περιαγγέλλω": ["2.85.3"], # vr prosperi-
        "περίειμι": ["1.30.3", "7.21.4"], # vr
        "περίπλοος2": ["2.97.1"], # actually periploos1
        "πλάγιος": ["4.35.4", "7.40.5"], # under plagion
        "πλέω": ["2.94.3", "6.51.3", "7.20.3"], # vr apo-, vr dia-, vr pros-
        "πλῆθος": ["8.92.9"], # vr?
        "πλησίος": ["2.4.5"], # struck from our text
        "πολεμέω": ["3.82.1"], # listed under polemow
        "πολεμικός": ["2.89.9"], # vr πολεμίων
        "πολέμιος": ["4.80.3"], # vr polemos
        "πόλεμος": ["6.6.1"], # not present
        "πόλις": ["1.24.3"], # vr
        "πολιτεία": ["6.17.2"], # vr politwn for politeiwn
        "Πολίχνη": ["7.4.6"],
        "πολίχνη": ["8.14.3", "8.23.6"], # these two are crosslisted, leaving them
        "ποτε": ["6.38.2", "6.38.4", "8.108.4", "6.104.2", "8.73.5"], # doesn't seem to appear here (vr?) third probably vr o(/te, 5th is o(/pote
        "ποτός": ["2.49.5"], # vr?
        "πρᾶξις": ["1.39.3"], # not only does this not exist, the accompanying quote doesn't either?!?!?
        "προαγορεύω": ["1.139.1"], # the one reference properly under prolegw
        "προανάγω": ["8.16.2"], # vr ἐξανήγετο
        "πρόειμι": ["5.70.1", "7.79.1"], # vr pros-, unclear
        "προεξάγω": ["7.6.2"], # vr proagw
        "προέχω": ["1.15.1", "3.32.1", "4.11.3"], # vr pros-
        "προΐσχω": ["3.68.1"], # ref
        "προπέμπω": ["7.3.1"], # vr pros-
        "προσάγω": ["2.97.3"], # vr
        "προσδέω2": ["6.8.3"], # prodew 1
        "πρόσειμι2": ["4.110.2", "4.47.3", "6.97.5"], # vr, eselqein maybe, but other possibilities, vr pro- x2
        "προσκάθημαι": ["8.76.5"], # ref
        "προσκαλέω": ["8.98.2"], # ref
        "προσκτάομαι": ["2.62.3"], # vr pro-
        "προσοικοδομέω": ["2.76.3"], # vr eis-
        "προσπίπτω": ["2.83.3"], # ref
        "προστασία": ["2.80.5"], # vr
        "προστίθημι": ["3.40.1"], # vr pro-
        "προσχωρέω": ["6.18.5"], # vr pro-
        "προτεραῖος": ["5.75.4", "7.51.2"], # vr προτέρᾳ
        "πρότερος": ["1.59.2", "7.19.5"], # vr prwtos x2
        "προφθάνω": ["7.73.1"], # vr prodialabontas
        "ῥώννυμι": ["6.17.8"], # seems to pretty clearly be from errw
        "σαφής": ["4.125.1"], # ref
        "στάδιον": ["6.1.2", "4.38.5"], # vr εἰκοσιστάδιος, different parse
        "στρατεύω": ["2.79.1", "2.80.5"], # vr epi-, vr cun-
        "στρατιά": ["1.27.1", "1.10.3", "5.60.6", "5.79.3", "7.55.1"], # vr strateia
        "στρατόπεδον": ["2.78.1"], # vr stratou=
        "συλλέγω": ["8.67.2"], # ref
        "συμβιβάζω": ["8.98.3"], # ref
        "συμμαχία": ["2.10.1", "5.33.2", "8.49.1"], # vr cummaxida x2, vr cunwmosia
        "συμμαχικός": ["8.7.1"], # ref
        "σύμμαχος": ["2.89.4"], # I can't find it
        "σύμπας": ["8.45.3"], # regular pantos
        "συμπίπτω": ["6.100.2"], # ref
        "συμπολεμέω": ["8.13.1"], # ref
        "συναιρέω": ["2.29.4", "8.24.5"], # refs
        "σύνειμι2": ["1.3.4", "3.111.2"], # cun-ec- x2
        "συνίστημι": ["1.90.2"], # vr
        "συνοικίζω": ["6.2.6"], # vr
        "συντάσσω": ["5.9.6"], # vr
        "συρράσσω": ["1.66.1"], # listed under surrhgnumi
        "σχολή": ["1.84.1", "2.75.6", "4.47.3", "7.15.2", "7.81.2"], # comparatives listed under sxolaios
        "τειχισμός": ["5.17.2"], # vr epi-
        "τεσσαράκοντα": ["3.30.4"], # ref
        "τετρακισχίλιοι": ["6.31.2"], # not really an occurence, two words are split by gar
        "τομή": ["1.93.5"], # ref
        "τόπος": ["6.54.4"], # ref
        "τοσόσδε": ["1.37.5"], # vr as two words
        "τοσοῦτος": ["8.76.5"], # vr toiouton
        "τότε": ["8.62.3", "7.60.5"], # vr pote, ref
        "τρεισκαίδεκα": ["3.69.1", "3.79.2", "8.88.1", "6.74.2"], # split into 3 words
        "τρέπω": ["7.49.2"], # ref
        "ὑπάγω": ["3.63.2"], # vr
        "ὑπεξέρχομαι": ["4.83.3"], # ref
        "ὑπερβολή": ["7.89.4"], # ref
        "ὑποκρίνομαι": ["7.44.5"], # vr apo-
        "ὑποτοπέω": ["5.116.1"], # vr upopteuw
        "ὑστερίζω": ["8.44.3"], # ref
        "φαίνω": ["8.44.3"], # vr epi-
        "φέρω": ["6.20.4"], # vr es-
        "φημί": ["7.6.1"], # unclear why this is here... did he mean 7.5.3?
        "φθείρω": ["2.91.1"], # vr dia-
        "φίλος": ["5.41.3"], # ref
        "φιλία": ["7.50.1"], # ref
        "φυλακή": ["6.101.5", "8.92.4"], # vr fulh x2
        "φύλαξ": ["2.24.1", "2.78.2", "4.90.4"], # vr x3
        "χράομαι": ["1.14.1", "5.58.4", "7.62.2", "8.40.2"], # under xraw x3, ref
        "χρῆμα": ["5.115.4"], # vr
        "χρόνος": ["8.91.1"], # ref
        "χῶρος": ["2.19.2"], # vr
        "ψηφίζω": ["4.88.1"], # ref
        "ψόφος": ["3.22.4"], # looks like this occurrence was deleted
        "ὡς": ["7.24.2"], # vr wsper
        "ὠφέλεια": ["1.141.7"], # not here
        "ὥσπερ": ["3.76.1", "8.34.1", "8.65.2"], # o(sper, ??, ref
        "ἀγνώς": ["3.94.5"], # wrong (Rusten)
        "ἀποστρέφω": ["4.80.1"], # VR (Rusten)
        "ἐπιφανής": ["8.87.4"], # VR (Rusten)
        "ἐμβολή": ["7.40.5"], # wrong (Rusten)
        "ἱεράομαι": ["5.1.1"], # betant is wrong  (Rusten)
        "νεωστί": ["1.7.1"], # wrong  (Rusten)
        "πῦρ": ["7.80.1", "7.80.3", "8.102.1"], # listed under neuter plurals
        "ἄκρος": ["1.105.3", "4.113.2"], # per JR, these make more sense as akron
        "Βοιωταρχέω": ["2.2.1", "4.91.1"],# per JR, should be from lowercase version
        "θεσμοφύλακες": ["5.47.9"], # we have singular version
        "Ἰσθμιάς": ["8.10.1"], # per JR, this should be parsed differently
        "παιδικόν": ["1.132.5"], # per JR, this should be parsed differently
        "σκηνέω": ["1.89.3", "2.52.3"], # tricky word, JR has different parses for both of these.
        "δεῖ": ["8.17.3", "8.25.1", "8.102.1"], # both Betant and Logeion correct, leaving this alone for now.
        "παραθαλασσίδιος": ["6.62.3"], # per JR, should be listed under different form
    }

    # get invalid links
    out = []
    for lem in lemmas:
        for mean in lemmas[lem].instanceMeanings:
            if not(mean in sectionLemmas):
                out.append("Bad Section \"%s\": [\"%s\"],  **" % (lem, mean))
            else:
                lemPresent = False
                lemAccountedFor = False
                if (lem in sectionLemmas[mean]):
                    lemPresent = True
                # ignore things that are dualed with erxomai
                if (lemmas[lem].betantLongDefinition == "true"):
                    if (lem in eimiMap and eimiMap[lem] in sectionLemmas[mean]):
                        lemPresent = True
                    if (lem in accounted and mean in accounted[lem]):
                        lemAccountedFor = True
                if not(lemPresent or lemAccountedFor):
                    out.append("Invalid Link \"%s\": [\"%s\"]," % (lem, mean))

    ERROR_LOG.append("Dictionary definition links that don't match the text:")
    ERROR_LOG.append("----------------")
    ERROR_LOG.append("\n".join(out))
    ERROR_LOG.append("================")

    for key in preppedTexts:
        fName = "results/prepTexts/" + key + ".txt"

        # create sections list
        utils.safeWrite(fName, re.sub(r'\s+', " ", preppedTexts[key]))

    # create sections list
    f = open('results/sections-list.ts', 'w')
    f.write("export const SECTIONS = " + json.dumps(sectionsList, indent=2))
    f.close()

    # print errors about lemmas from perseus's text analysis not in our dictionary
    ERROR_LOG.append("Lemmas in Perseus's analysis of the text but not our dictionary:")
    ERROR_LOG.append("----------------")
    for key in sorted(list(noMatchLemmas.keys()), key=simplifyUnicode.unaccented):
        s = "%s:  ['%s']" % (key, "', '".join(noMatchLemmas[key]))
        ERROR_LOG.append(s)
    ERROR_LOG.append("================")

    # print errors about sections that don't exist
    ERROR_LOG.append("Long Definition Sections that don't exist:")
    ERROR_LOG.append("----------------")
    for key in lemmas:
        l = lemmas[key]
        bads = parseDefinition.getBadSections(l, allSections)
        if (len(bads) > 0):
            s = "%s: %s" % (key, ", ".join(bads))
            ERROR_LOG.append(s)
    ERROR_LOG.append("================")



    conn = sqlite3.connect('results/lexicon_database.db')
    c = conn.cursor()

    # Create lemma table
    c.execute('DROP TABLE IF EXISTS lemmata')
    c.execute('''CREATE TABLE lemmata
                 (lemma text,
                  short_def text,
                  has_long_def text,
                  old_long_def text,
                  author_id text,
                  custom_author text,
                  long_def_raw text,
                  long_def text,
                  part_of_speech text,
                  semantic_group text,
                  root text,
                  compounds text,
                  frequency_all integer,
                  has_key_passage integer,
                  key_passage_location text,
                  key_passage_text text,
                  key_passage_translation text,
                  has_illustration integer,
                  illustration_source text,
                  illustration_alt text,
                  illustration_caption text,
                  bibliography_text text)''')
    c.execute('CREATE INDEX idx_lemmata_l ON lemmata(lemma)')


    nonOccuringLemmas = []
    for key in lemmas:
        l = lemmas[key]
        command = l.getSQLiteInsertCommand("lemmata")
        if l.frequency["all"] == 0:
            nonOccuringLemmas.append(key)
        #print(command)
        c.execute(command)

    ERROR_LOG.append("Lemmas that appear 0 times:")
    ERROR_LOG.append("----------------")
    ERROR_LOG.append("\n".join(nonOccuringLemmas))
    ERROR_LOG.append("================")


    # Create table for searching
    c.execute('DROP TABLE IF EXISTS search_lemmata')
    c.execute('''CREATE TABLE search_lemmata
                 (lemma text,
                  alias text,
                  search_text text)''')
    c.execute('CREATE INDEX idx_search_lemmata_st ON search_lemmata(search_text)')

    for key in lemmas:
        commands = lemmas[key].getSQLiteSearchCommands("search_lemmata")
        for command in commands:
            #print(command)
            c.execute(command)

    # Add Aliases (e.g. πλείων -> πολύς)
    c.execute('DROP TABLE IF EXISTS aliases')
    c.execute('''CREATE TABLE aliases
                 (alias text,
                  lemma text)''')
    c.execute('CREATE INDEX idx_aliases_lemma ON aliases(lemma)')
    c.execute('CREATE INDEX idx_aliases_alias ON aliases(alias)')

    aliasWorkbook = load_workbook('input/aliases.xlsx', read_only=True)
    aliasSheet = aliasWorkbook['Sheet1']
    firstRow = True

    for row in aliasSheet:
        # Skip first row
        if (firstRow):
            firstRow = False
            continue

        # Skip empty rows
        if (row[0].value == None):
            continue


        alias = row[0].value
        aliasLemma = row[1].value

        command = "INSERT INTO aliases VALUES (\'%s\', \'%s\')" % (sqliteEscapeString(alias), sqliteEscapeString(aliasLemma));
        c.execute(command)

        searchLemma = simplifyUnicode.searchVersion(alias)
        unaccentedLemma = simplifyUnicode.unaccented(alias)
        betacodeLemma = simplifyUnicode.betacode(alias)
        betacodeUnaccentedLemma = simplifyUnicode.unaccentedBetacode(alias)
        latinApproximationLemma = simplifyUnicode.latinApproximation(alias)

        searches = [
            searchLemma,
            unaccentedLemma,
            betacodeLemma,
            betacodeUnaccentedLemma,
            latinApproximationLemma
        ]

        # final sigma
        if (alias[-1] == "ς"):
            searches.append(searchLemma[:-1] + "σ")
            searches.append(unaccentedLemma[:-1] + "σ")

        for search in searches:
            s = []
            s.append("INSERT INTO search_lemmata VALUES (")
            s.append("\'\', \'%s\', " % sqliteEscapeString(alias))

            # search information
            s.append("\'%s\'" % sqliteEscapeString(search))
            s.append(")")
            command = "".join(s)
            c.execute(command)



    # Create compounds table
    c.execute('DROP TABLE IF EXISTS compounds')
    c.execute('''CREATE TABLE compounds
                (compound_index integer,
                 compound text,
                 description text,
                 lemma_in_dict integer )''')

    # Insert data
    index = 0
    for comp in allCompounds:
        in_dict = 1
        if not(comp in lemmas):
            in_dict = 0
        allCompounds[comp] = index
        command = "INSERT INTO compounds VALUES (%d,'%s','', %d)" % (index, comp, in_dict)
        c.execute(command)
        index += 1

    # Create compound-lemma link
    c.execute('DROP TABLE IF EXISTS compound_lemma_link')
    c.execute('''CREATE TABLE compound_lemma_link
            (compound text,
             lemma text)''')
    c.execute('CREATE INDEX idx_compound_lemma_link_c ON compound_lemma_link(compound)')

    for key in lemmas:
        l = lemmas[key]
        compounds = l.compoundParts
        for comp in compounds:
            command = "INSERT INTO compound_lemma_link VALUES ('%s','%s')" % (comp, key)
            c.execute(command)

    #====

    # Create stem types table
    c.execute('DROP TABLE IF EXISTS roots')
    c.execute('''CREATE TABLE roots
                 (root_index integer,
                  root text,
                  description text,
                  lemma_in_dict integer)''')

    # Insert data
    index = 0
    for st in stemTypes:
        in_dict = 1;
        if not(st in lemmas):
            in_dict = 0;
        command = "INSERT INTO roots VALUES (%d,'%s','', %d)" % (index, st, in_dict)
        c.execute(command)
        index += 1

    # Create stem-lemma link
    c.execute('DROP TABLE IF EXISTS root_lemma_link')
    c.execute('''CREATE TABLE root_lemma_link
            (stem text,
             lemma text)''')
    c.execute('CREATE INDEX idx_root_lemma_link_s ON root_lemma_link(stem)')

    for key in lemmas:
        l = lemmas[key]
        sts = l.stemType
        for st in sts:
            command = "INSERT INTO root_lemma_link VALUES ('%s','%s')" % (st, key)
            c.execute(command)
    #====


    # Create semantic groups table
    c.execute('DROP TABLE IF EXISTS semantic_groups')
    c.execute('''CREATE TABLE semantic_groups
                 (group_index integer,
                  group_name text,
                  label_class text,
                  description text)''')

    # Create semantic group-lemma link
    c.execute('DROP TABLE IF EXISTS semantic_lemma_link')
    c.execute('''CREATE TABLE semantic_lemma_link
          (semantic_group integer,
           lemma text)''')
    c.execute('CREATE INDEX idx_semantic_lemma_link_s ON semantic_lemma_link(semantic_group)')


    semanticGroupList = []
    for group in semanticGroups:
        semanticGroupList.append(group)


    semanticGroupList = sorted(semanticGroupList)

    for i in range(len(semanticGroupList)):
        groupIndex = i
        groupName = semanticGroupList[i]
        groupClass = createSemanticGroupBadgeClass(groupName)
        groupDesc = groupName
        semanticIndices[groupName] = groupIndex
        c.execute("INSERT INTO semantic_groups VALUES (%d,'%s','%s','%s')" % (groupIndex, groupName, groupClass, groupDesc))


    for key in lemmas:
      l = lemmas[key]
      sgs = l.semanticGroup
      for sg in sgs:
          sg_index = semanticIndices[sg]
          command = "INSERT INTO semantic_lemma_link VALUES ('%s','%s')" % (sg_index, key)
          c.execute(command)


    # text_storage table
    c.execute('DROP TABLE IF EXISTS text_storage')
    c.execute('''CREATE TABLE text_storage
         (sequence_index integer,
          token_index integer,
          token text,
          book integer,
          chapter integer,
          section integer,
          word_index integer,
          true_word_index integer)''')
    c.execute('CREATE UNIQUE INDEX idx_text_storage_si ON text_storage(sequence_index)')
    c.execute('CREATE INDEX idx_text_storage_ti ON text_storage(token_index)')

    # Insert data
    for i in textStorageInputs:
        c.execute(i)


    # instance_information
    c.execute('DROP TABLE IF EXISTS instance_information')
    c.execute('''CREATE TABLE instance_information
         (token_index integer,
          token text,
          book integer,
          chapter integer,
          section integer,
          word_index integer,
          true_word_index integer,
          section_code text,
          lemma text,
          lemma_meaning text,
          context_type integer)''')
    c.execute('CREATE UNIQUE INDEX idx_instance_information_ti ON instance_information(token_index)')
    c.execute('CREATE INDEX idx_instance_information_l ON instance_information(lemma)')

    # Insert data
    for i in instanceInformationInputs:
      c.execute(i)

    # section_contexts
    c.execute('DROP TABLE IF EXISTS section_contexts')
    c.execute('''CREATE TABLE section_contexts
         (token_index integer,
          book integer,
          chapter integer,
          section integer,
          word_index integer,
          true_word_index integer,
          context_type integer)''')
    c.execute('CREATE UNIQUE INDEX idx_section_contexts_ti ON section_contexts(token_index)')

    # section_code context_type
    for i in sectionContextInputs:
        c.execute(i)

    # save and close database
    conn.commit()
    conn.close()


# ==============================================================================
# ================================== Run Code  =================================
# ==============================================================================
lems = getLemmaInfo()

missing = 0
betantHasWeDont = []
usedTwice = []
for key in betantUsed:
    if len(betantUsed[key]) == 0:
        #betantHasWeDont.append("%s, %s")
        betantHasWeDont.append("%s (%s)" % (betant[key][1][0]["text"]["lem"], key))
        missing += 1
    elif len(betantUsed[key]) > 1:
        usedTwice.append("%s (%s)" % (key, ", ".join(betantUsed[key])))

ERROR_LOG.append("Lemmas that appear twice in the Excel File:")
ERROR_LOG.append("----------------")
ERROR_LOG.append("\n".join(usedTwice))
ERROR_LOG.append("================")

out = ["Total missing: %d" % missing]
out.extend(betantHasWeDont)
utils.safeWrite("results/betantHasWeDont.txt", "\n".join(out))
utils.safeWrite("results/weHaveBetantDoesnt.txt", "\n".join(betantMissing))
#


workbook = load_workbook('input/contexts.xlsx', read_only=True)
main_sheet = workbook['Sheet1']

createSQLiteDatabase(lems)

f = open('results/errors.txt', 'w')
f.write("\n".join(ERROR_LOG))
f.close()
